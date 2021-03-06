# -*- coding: utf-8 -*-  
import openpyxl
import datetime
import sys
import requests
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Alignment
import json as json
import os
basedir = os.path.abspath(os.path.dirname(__file__))

koef = 2
defwidth = 17
defheight = 15
colours = {'blue':'ADD8E6', 'red':'FF0043', 'green':'28FF3E', 
          'orange':'FFAD42','white':'FFFFFF', 'yellow':'FFFF01', 'grey':'BC8D8D'}
print(sys.argv)
days = ['Понедельн.', 'Вторник', 'Среда', 'Четверг', 'Пяница', 'Суббота']
months = ['Янв.','Фев.','Мар.','Апр.','Мая','Июн.','Июл.','Авг.','Сен.','Окт.','Ноя.','Дек']
r = requests.get("http://localhost/var/www/html/Raspisator/API/getProject.php?project_id="+sys.argv[1]+"&return_school_data=0")
r = json.loads(r.text)
if r['project_data'] == None:
     sys.exit()
s_id = r['school_id']
a = json.loads(r['project_data'])
lpd = int(r['lessons_per_day'])
sinfo = requests.get("http://localhost/var/www/html/Raspisator/API/getSchoolData.php?school_id="+s_id+"")
sinfo = json.loads(sinfo.text)
startdate = datetime.datetime.strptime(r['start'], '%Y-%m-%d')
finishdate = datetime.datetime.strptime(r['finish'], '%Y-%m-%d')
print(startdate.weekday())
les = requests.get("http://localhost/var/www/html/Raspisator/API/getLessons.php?project_id="+sys.argv[1])
les = json.loads(les.text)
les = {i['id'] : i for i in les}
print(sinfo)
grades = sinfo['grades']
teachers = sinfo['teachers']
subjects = sinfo['subjects']

wb = openpyxl.workbook.Workbook()
ws = wb.active
ws.merge_cells('A'+str(koef-1)+':'+str(unichr(ord('A')+len(grades)+1))+str(koef-1))
ws['A'+str(koef-1)] = "Расписание на " + str(startdate.date().day)+" "+months[(startdate).date().month-1] + " - " + str(finishdate.date().day)+" "+months[finishdate.date().month-1]
ws['A'+str(koef-1)].font = Font(color=colors.BLACK, bold=True)
ws['A'+str(koef-1)].alignment = Alignment(wrap_text = True, horizontal = 'center')

for i in range(len(a['grades'])):
     #классы
     ws[str(unichr(ord('A')+i+2))+str(koef)] = grades[a['grades'][i]]['grade_number'] + grades[a['grades'][i]]['grade_name']
     t = ws[str(unichr(ord('A')+i+2))+str(koef)]
     t.font = Font(color=colors.BLACK, bold=True)
     ws.column_dimensions[str(unichr(ord('A')+i+2))].width=defwidth
     #t.fill = PatternFill(bgColor=colors.RED, fill_type = "solid")
ws.column_dimensions['A'].width = 9

koef+=1

for i in range(len(a['table']['table'])):
     for j in range(len(a['table']['table'][i])):
          if(i % lpd == 0 and j == 0):
               #номера
               for k in range(lpd):
                    ws['B'+str(i+k+koef+i/lpd)].alignment = Alignment(wrap_text = True, horizontal = 'center')
                    ws['B'+str(i+k+koef+i/lpd)] = k+1
                    ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width, len(str(k+1))+1)
               
               #серая полоска
               if(i != 0):
                    ws.merge_cells('A'+str(i+koef-1+i//lpd)+':'+str(unichr(ord('A')+len(grades)+1))+str(i+koef-1+i//lpd))
                    ws['A'+str(i+koef-1+i//lpd)].fill = PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type = "solid")     
               
               #дата и день недели
               ws.merge_cells('A'+str(i+koef+i//lpd)+':'+'A'+str(i+koef-1+lpd+i//lpd))               
               ws['A'+str(i+koef+i//lpd)].alignment = Alignment(wrap_text = True, horizontal = 'center', vertical = 'center')
               ws['A'+str(i+koef+i//lpd)] = days[((startdate+datetime.timedelta(days=i//lpd)).weekday())%6]+"\n"+str((startdate+datetime.timedelta(days=i//lpd)).date().day)+" "+months[(startdate+datetime.timedelta(days=i//lpd)).date().month-1]
          
          #сами пары
          if(a['table']['table'][i][j] != -1):
               temp = a['lessons'][a['table']['table'][i][j]]['db_id']
               st = ""
               st = st + subjects[les[temp]['subject_id']]['name'] + ' (' + teachers[les[temp]['teacher_id']]['name'] + ')'
               ws[str(unichr(ord('A')+j+2))+str(i+koef+i//lpd)].alignment = Alignment(wrap_text = True, vertical = 'center')
               tempc = a['lessons'][a['table']['table'][i][j]]['color'][1:]
               ws[str(unichr(ord('A')+j+2))+str(i+koef+i//lpd)].fill=PatternFill(start_color=tempc,fill_type= "solid")
               ws[str(unichr(ord('A')+j+2))+str(i+koef+i//lpd)] = st 
               ws.row_dimensions[i+2+i//lpd].height = max(ws.row_dimensions[i+2+i//lpd].height, defheight*(len(st)+len(st)-1)//defwidth)
               

file_path = os.path.join(basedir, 'excell/'+str(r['id'])+".xlsx")
print(ws['A2'])
wb.save(open(file_path,'w'))
